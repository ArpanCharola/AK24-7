import io
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_

from app.core.database import get_db
from app.core.auth import get_current_user
from app.models.user import User
from app.models.discovered_job import DiscoveredJob
from app.models.job_application import JobApplication
from app.workers.tasks import run_application_task, _detect_portal

router = APIRouter()


class DiscoveredJobResponse(BaseModel):
    id: int
    job_url: str
    title: Optional[str]
    company: Optional[str]
    location: Optional[str]
    source: str
    match_score: Optional[int]
    match_reason: Optional[str]
    status: str
    discovered_at: datetime
    work_arrangement: Optional[str]
    posted_at: Optional[datetime]
    contact_email: Optional[str] = None
    contact_linkedin: Optional[str] = None
    contact_name: Optional[str] = None
    contact_source: Optional[str] = None
    contact_enriched_at: Optional[datetime] = None

    model_config = ConfigDict(from_attributes=True)


@router.get("/discovered-jobs", response_model=list[DiscoveredJobResponse])
async def list_discovered(
    status: Optional[str] = Query(None),
    posted_within_days: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(DiscoveredJob).where(DiscoveredJob.user_id == current_user.id)
    if status:
        q = q.where(DiscoveredJob.status == status)
    if posted_within_days:
        cutoff = datetime.now(timezone.utc) - timedelta(days=posted_within_days)
        # Keep rows with an unknown posting date — a NULL posted_at means the
        # source didn't expose one, not that the job is old; dropping them would
        # silently hide jobs from the list/export.
        q = q.where(or_(DiscoveredJob.posted_at >= cutoff, DiscoveredJob.posted_at.is_(None)))
    q = q.order_by(
        DiscoveredJob.match_score.desc().nulls_last(),
        DiscoveredJob.discovered_at.desc(),
    )
    result = await db.execute(q)
    return result.scalars().all()


@router.post("/discovered-jobs/{job_id}/queue", status_code=201)
async def queue_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(DiscoveredJob).where(
            DiscoveredJob.id == job_id,
            DiscoveredJob.user_id == current_user.id,
        )
    )
    discovered = result.scalar_one_or_none()
    if not discovered:
        raise HTTPException(status_code=404, detail="Discovered job not found")

    # Dedupe — covers both manually-queued and auto-queued paths
    existing_app = await db.execute(
        select(JobApplication).where(
            JobApplication.user_id == current_user.id,
            JobApplication.job_url == discovered.job_url,
        )
    )
    if existing_app.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Application already exists for this URL")

    application = JobApplication(
        user_id=current_user.id,
        job_url=discovered.job_url,
        job_title=discovered.title,
        company=discovered.company,
        job_description=discovered.job_description,
        portal_type=discovered.source,
        queued_by="manual",
    )
    db.add(application)
    discovered.status = "queued"
    await db.commit()
    await db.refresh(application)

    # Persist celery_task_id BEFORE dispatch so the worker never sees the row
    # without its task id.
    task_id = uuid.uuid4().hex
    application.celery_task_id = task_id
    await db.commit()
    run_application_task.apply_async(args=[application.id, current_user.id], task_id=task_id)

    return {"status": "queued", "application_id": application.id}


class BulkQueueRequest(BaseModel):
    ids: list[int]


@router.post("/discovered-jobs/bulk-queue", status_code=201)
async def bulk_queue_jobs(
    body: BulkQueueRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Queue many discovered jobs for apply in one request.

    Mirrors the single /queue endpoint: for each owned, still-queueable job that
    hasn't already been applied, create a manual JobApplication, mark the
    discovered row "queued", and dispatch the apply task. Returns a per-job
    summary so the UI can report partial results.
    """
    if not body.ids:
        return {"queued": [], "skipped": []}

    rows = await db.execute(
        select(DiscoveredJob).where(
            DiscoveredJob.id.in_(body.ids),
            DiscoveredJob.user_id == current_user.id,
        )
    )
    jobs = {j.id: j for j in rows.scalars().all()}

    # Dedupe against everything this user has already applied to (manual + auto).
    applied_rows = await db.execute(
        select(JobApplication.job_url).where(JobApplication.user_id == current_user.id)
    )
    already_applied = {r[0] for r in applied_rows.all()}

    queued: list[dict] = []
    skipped: list[dict] = []
    pending: list[int] = []  # application ids to dispatch after commit
    seen_urls: set[str] = set()

    for jid in body.ids:
        job = jobs.get(jid)
        if not job:
            skipped.append({"id": jid, "reason": "not found"})
            continue
        if job.status not in ("discovered", "auto_queued"):
            skipped.append({"id": jid, "reason": f"status is '{job.status}'"})
            continue
        if job.job_url in already_applied or job.job_url in seen_urls:
            skipped.append({"id": jid, "reason": "already applied"})
            continue

        application = JobApplication(
            user_id=current_user.id,
            job_url=job.job_url,
            job_title=job.title,
            company=job.company,
            job_description=job.job_description,
            portal_type=job.source or _detect_portal(job.job_url),
            queued_by="manual",
        )
        db.add(application)
        job.status = "queued"
        await db.flush()  # assign application.id without ending the transaction
        task_id = uuid.uuid4().hex
        application.celery_task_id = task_id  # committed with the row below
        seen_urls.add(job.job_url)
        pending.append((application.id, task_id))
        queued.append({"id": jid, "application_id": application.id})

    # Commit all rows (incl. celery_task_id) BEFORE dispatch so the worker always
    # sees a committed application with its task id.
    await db.commit()

    for app_id, task_id in pending:
        run_application_task.apply_async(args=[app_id, current_user.id], task_id=task_id)

    return {"queued": queued, "skipped": skipped}


class FindContactResponse(BaseModel):
    contact_email: Optional[str] = None
    contact_linkedin: Optional[str] = None
    contact_name: Optional[str] = None
    contact_source: Optional[str] = None
    enriched_at: Optional[datetime] = None


@router.post("/discovered-jobs/{job_id}/find-contact", response_model=FindContactResponse)
async def find_contact_for_job(
    job_id: int,
    use_apify: bool = Query(False, description="Spend Apify credits to scrape company site"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Run the full contact-enrichment cascade for a single discovered job.

    Free tiers (regex + LLM-named-recruiter + generic guess) always run.
    Apify scrape runs only when `use_apify=true` AND APIFY_TOKEN is configured.
    Updates the row in place and returns whatever was found.
    """
    from app.services.contact_finder import find_contact

    result = await db.execute(
        select(DiscoveredJob).where(
            DiscoveredJob.id == job_id,
            DiscoveredJob.user_id == current_user.id,
        )
    )
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Discovered job not found")

    job_dict = {
        "job_description": job.job_description,
        "company": job.company,
        "job_url": job.job_url,
    }
    found = await find_contact(job_dict, use_apify=use_apify)

    now = datetime.now(timezone.utc)
    # Only overwrite a non-null field if we found something better. We treat
    # any new email/linkedin from a non-`generic_unverified` source as better.
    if found:
        new_source = found.get("source")
        promote = new_source and new_source != "generic_unverified"
        if found.get("email") and (promote or not job.contact_email):
            job.contact_email = found["email"]
        if found.get("linkedin") and (promote or not job.contact_linkedin):
            job.contact_linkedin = found["linkedin"]
        if found.get("name") and (promote or not job.contact_name):
            job.contact_name = found["name"]
        if new_source:
            job.contact_source = new_source
    job.contact_enriched_at = now
    await db.commit()

    return FindContactResponse(
        contact_email=job.contact_email,
        contact_linkedin=job.contact_linkedin,
        contact_name=job.contact_name,
        contact_source=job.contact_source,
        enriched_at=job.contact_enriched_at,
    )


@router.post("/discovered-jobs/{job_id}/skip")
async def skip_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(DiscoveredJob).where(
            DiscoveredJob.id == job_id,
            DiscoveredJob.user_id == current_user.id,
        )
    )
    discovered = result.scalar_one_or_none()
    if not discovered:
        raise HTTPException(status_code=404, detail="Discovered job not found")
    discovered.status = "skipped"
    await db.commit()
    return {"status": "skipped"}


@router.delete("/discovered-jobs/{job_id}", status_code=204)
async def delete_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(DiscoveredJob).where(
            DiscoveredJob.id == job_id,
            DiscoveredJob.user_id == current_user.id,
        )
    )
    discovered = result.scalar_one_or_none()
    if not discovered:
        raise HTTPException(status_code=404, detail="Discovered job not found")
    await db.delete(discovered)
    await db.commit()


@router.get("/discovered-jobs/export")
async def export_discovered_jobs(
    status: Optional[str] = Query(None),
    posted_within_days: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download discovered jobs as an Excel (.xlsx) file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = select(DiscoveredJob).where(DiscoveredJob.user_id == current_user.id)
    if status:
        q = q.where(DiscoveredJob.status == status)
    if posted_within_days:
        cutoff = datetime.now(timezone.utc) - timedelta(days=posted_within_days)
        # Keep rows with an unknown posting date — a NULL posted_at means the
        # source didn't expose one, not that the job is old; dropping them would
        # silently hide jobs from the list/export.
        q = q.where(or_(DiscoveredJob.posted_at >= cutoff, DiscoveredJob.posted_at.is_(None)))
    q = q.order_by(
        DiscoveredJob.match_score.desc().nulls_last(),
        DiscoveredJob.discovered_at.desc(),
    )
    result = await db.execute(q)
    jobs = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discovered Jobs"

    headers = ["Title", "Company", "Location", "Source", "Work Arrangement",
               "Match Score", "Status", "Posted At", "Discovered At", "Job URL", "Match Reason"]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, job in enumerate(jobs, 2):
        ws.cell(row=row, column=1, value=job.title or "")
        ws.cell(row=row, column=2, value=job.company or "")
        ws.cell(row=row, column=3, value=job.location or "")
        ws.cell(row=row, column=4, value=job.source or "")
        ws.cell(row=row, column=5, value=job.work_arrangement or "")
        ws.cell(row=row, column=6, value=job.match_score)
        ws.cell(row=row, column=7, value=job.status or "")
        ws.cell(row=row, column=8, value=job.posted_at.strftime("%Y-%m-%d") if job.posted_at else "")
        ws.cell(row=row, column=9, value=job.discovered_at.strftime("%Y-%m-%d %H:%M") if job.discovered_at else "")
        url_cell = ws.cell(row=row, column=10, value=job.job_url or "")
        if job.job_url:
            url_cell.hyperlink = job.job_url
            url_cell.font = Font(color="4F46E5", underline="single")
        ws.cell(row=row, column=11, value=job.match_reason or "")

    col_widths = [35, 22, 20, 14, 18, 12, 12, 14, 18, 55, 60]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"discovered_jobs_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class BulkDeleteRequest(BaseModel):
    ids: list[int]


@router.post("/discovered-jobs/bulk-delete", status_code=204)
async def bulk_delete_jobs(
    body: BulkDeleteRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(DiscoveredJob).where(
            DiscoveredJob.id.in_(body.ids),
            DiscoveredJob.user_id == current_user.id,
        )
    )
    jobs = result.scalars().all()
    for job in jobs:
        await db.delete(job)
    await db.commit()
